import sys
import cv2
import os
import numpy as np
import torch
from PyQt5.QtWidgets import (
    QApplication,
    QLabel,
    QVBoxLayout,
    QWidget,
    QPushButton,
    QInputDialog
)
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtCore import QTimer, QDateTime
from facenet_pytorch import InceptionResnetV1
from scipy.spatial.distance import cosine
import pickle
from openpyxl import Workbook, load_workbook
import time

# ===============================
# Model Setup
# ===============================
device = torch.device("cpu")  # CPU saja di RPi
resnet = InceptionResnetV1(pretrained="vggface2").eval().to(device)

# Haar Cascade
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

# Embedding DB
db_path = "face_embeddings_model.pkl"
base_dataset_dir = "Dataset_Face"

def load_embeddings():
    if os.path.exists(db_path):
        with open(db_path, "rb") as f:
            model_data = pickle.load(f)
        embeddings = model_data["embeddings"]
        names = model_data["names"]
        embeddings = embeddings / np.linalg.norm(embeddings, axis=1, keepdims=True)
    else:
        embeddings = np.empty((0,512))
        names = np.array([])
    return embeddings, names

db_embeddings, db_names = load_embeddings()

# ===============================
# Main App
# ===============================
class FaceRecogApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Face Recognition (Lightweight)")

        self.video_label = QLabel()
        self.status_label = QLabel("Status: Initializing...")

        self.add_button = QPushButton("Add Data")
        self.add_button.clicked.connect(self.add_data)

        self.refresh_button = QPushButton("Refresh Model")
        self.refresh_button.clicked.connect(self.refresh_model)

        layout = QVBoxLayout()
        layout.addWidget(self.video_label)
        layout.addWidget(self.status_label)
        layout.addWidget(self.add_button)
        layout.addWidget(self.refresh_button)
        self.setLayout(layout)

        self.cap = cv2.VideoCapture(0)
        if self.cap.isOpened():
            self.status_label.setText("Status: Camera Started")
        else:
            self.status_label.setText("Status: Failed to open camera")

        # Logging setup
        self.log_path = "face_recognition_log.xlsx"
        if not os.path.exists(self.log_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["Timestamp", "Identity", "Image Path"])
            wb.save(self.log_path)

        self.capture_dir = "capture_logs"
        os.makedirs(self.capture_dir, exist_ok=True)

        self.last_logged_time = 0
        self.last_identity = None

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_frame)
        self.timer.start(200)  # Update every 200 ms

    def update_frame(self):
        ret, frame = self.cap.read()
        if not ret:
            return

        frame = cv2.resize(frame, (320, 240))
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5)

        recognized = False

        for (x, y, w, h) in faces:
            cv2.rectangle(frame, (x, y), (x+w, y+h), (0,255,0), 2)
            face_roi = frame[y:y+h, x:x+w]
            if face_roi.size == 0:
                continue

            face_rgb = cv2.cvtColor(face_roi, cv2.COLOR_BGR2RGB)
            face_resized = cv2.resize(face_rgb, (160,160))
            face_tensor = torch.tensor(face_resized).permute(2,0,1).float()/255.0
            face_tensor = (face_tensor - 0.5)/0.5
            face_tensor = face_tensor.unsqueeze(0).to(device)

            with torch.no_grad():
                embedding = resnet(face_tensor).cpu().numpy()[0]
                embedding = embedding / np.linalg.norm(embedding)

            max_similarity = -1
            identity = "Unknown"
            for db_emb, name in zip(db_embeddings, db_names):
                similarity = 1 - cosine(embedding, db_emb)
                if similarity > max_similarity:
                    max_similarity = similarity
                    identity = name

            if max_similarity < 0.7:
                identity = "Unknown"

            label = f"{identity} ({max_similarity:.2f})"
            cv2.putText(frame, label, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,0), 2)

            # Log if recognized
            if identity != "Unknown":
                recognized = True
                if identity != self.last_identity or time.time() - self.last_logged_time > 10:
                    self.log_detection(identity, frame)
                    self.last_logged_time = time.time()
                    self.last_identity = identity

        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        qt_image = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
        self.video_label.setPixmap(QPixmap.fromImage(qt_image))

    def log_detection(self, identity, frame):
        timestamp = QDateTime.currentDateTime().toString("yyyy-MM-dd_HH-mm-ss")
        filename = f"{identity}_{timestamp}.jpg"
        filepath = os.path.join(self.capture_dir, filename)

        cv2.imwrite(filepath, frame)

        wb = load_workbook(self.log_path)
        ws = wb.active
        ws.append([timestamp.replace("_", " "), identity, filepath])
        wb.save(self.log_path)

        self.status_label.setText(f"Logged: {identity}")

    def add_data(self):
        name, ok = QInputDialog.getText(self, "Add New Face", "Enter Name:")
        if not ok or not name:
            return

        self.status_label.setText("Recording 5 seconds...")
        QApplication.processEvents()

        save_dir = os.path.join(base_dataset_dir, name)
        os.makedirs(save_dir, exist_ok=True)

        next_idx = 0
        start_time = QDateTime.currentDateTime()
        end_time = start_time.addSecs(5)

        while QDateTime.currentDateTime() < end_time:
            ret, frame = self.cap.read()
            if not ret:
                continue

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5)

            for (x, y, w, h) in faces:
                face_roi = frame[y:y+h, x:x+w]
                if face_roi.size == 0:
                    continue

                filename = os.path.join(save_dir, f"{name}_{next_idx:03d}.jpg")
                next_idx += 1
                cv2.imwrite(filename, face_roi)

            QApplication.processEvents()

        self.status_label.setText("Recording done. Refresh model to update.")

    def refresh_model(self):
        self.status_label.setText("Refreshing model...")
        QApplication.processEvents()

        all_embeddings = []
        all_names = []

        for person in os.listdir(base_dataset_dir):
            person_dir = os.path.join(base_dataset_dir, person)
            if not os.path.isdir(person_dir):
                continue

            person_embeddings = []

            for file in os.listdir(person_dir):
                if not file.lower().endswith(".jpg"):
                    continue
                img_path = os.path.join(person_dir, file)
                img = cv2.imread(img_path)
                if img is None:
                    continue

                face_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                face_resized = cv2.resize(face_rgb, (160,160))
                face_tensor = torch.tensor(face_resized).permute(2,0,1).float()/255.0
                face_tensor = (face_tensor - 0.5)/0.5
                face_tensor = face_tensor.unsqueeze(0).to(device)

                with torch.no_grad():
                    emb = resnet(face_tensor).cpu().numpy()[0]
                    emb = emb / np.linalg.norm(emb)
                    person_embeddings.append(emb)

            if person_embeddings:
                mean_emb = np.mean(person_embeddings, axis=0)
                all_embeddings.append(mean_emb)
                all_names.append(person)

        if all_embeddings:
            global db_embeddings, db_names
            db_embeddings = np.vstack(all_embeddings)
            db_names = np.array(all_names)

            with open(db_path, "wb") as f:
                pickle.dump({"embeddings": db_embeddings, "names": db_names}, f)

            self.status_label.setText("Model refreshed successfully.")
        else:
            self.status_label.setText("No embeddings found.")

    def closeEvent(self, event):
        if self.cap:
            self.cap.release()
        event.accept()

# ===============================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FaceRecogApp()
    window.show()
    sys.exit(app.exec_())
