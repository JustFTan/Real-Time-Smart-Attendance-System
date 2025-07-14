import sys
import serial
import cv2
import os
import numpy as np
import torch
from PyQt5.QtWidgets import (
    QApplication,
    QLabel,
    QVBoxLayout,
    QHBoxLayout,
    QWidget,
    QPushButton,
    QInputDialog,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView
)
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtCore import QTimer, QDateTime
from facenet_pytorch import InceptionResnetV1
from scipy.spatial.distance import cosine
import pickle
import mediapipe as mp
from openpyxl import Workbook, load_workbook
import time

# Load FaceNet model
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
resnet = InceptionResnetV1(pretrained="vggface2").eval().to(device)

# Haar Cascade for face detection
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

# Mediapipe FaceMesh
mp_face_mesh = mp.solutions.face_mesh
face_mesh = mp_face_mesh.FaceMesh(refine_landmarks=True)

# EAR calculation
LEFT_EYE = [33, 160, 158, 133, 153, 144]
RIGHT_EYE = [362, 385, 387, 263, 373, 380]
EAR_THRESHOLD = 0.2

def get_ear(landmarks, eye_indices, w, h):
    p = []
    for idx in eye_indices:
        lm = landmarks[idx]
        p.append(np.array([lm.x * w, lm.y * h]))
    A = np.linalg.norm(p[1] - p[5])
    B = np.linalg.norm(p[2] - p[4])
    C = np.linalg.norm(p[0] - p[3])
    ear = (A + B) / (2.0 * C)
    return ear

# Embedding database
db_path = "face_embeddings_model.pkl"
base_dataset_dir = r"C:/Users/justf/Downloads/AI-Camera-Door-Lock-System-main/AI-Camera-Door-Lock-System-main/Dataset_Face"

def load_embeddings():
    if os.path.exists(db_path):
        with open(db_path, "rb") as f:
            model_data = pickle.load(f)
        embeddings = model_data["embeddings"]
        names = model_data["names"]
        embeddings = embeddings / np.linalg.norm(embeddings, axis=1, keepdims=True)
    else:
        embeddings = np.empty((0,512))
        names = np.array([])
    return embeddings, names

db_embeddings, db_names = load_embeddings()

# Main App
class FaceRecogApp(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Face Recognition App")
        self.resize(1200, 600)

        self.video_label = QLabel()
        self.status_label = QLabel("Status: Initializing...")

        self.add_button = QPushButton("Add Data")
        self.add_button.clicked.connect(self.add_data)

        self.refresh_button = QPushButton("Refresh Model")
        self.refresh_button.clicked.connect(self.refresh_model)

        self.list_button = QPushButton("List Daftar Orang")
        self.list_button.clicked.connect(self.list_people)

        self.delete_button = QPushButton("Hapus Data Seseorang")
        self.delete_button.clicked.connect(self.delete_person)

        self.export_button = QPushButton("Export Log ke CSV")
        self.export_button.clicked.connect(self.export_log_csv)

        # Log Table
        self.log_table = QTableWidget()
        self.log_table.setColumnCount(3)
        self.log_table.setHorizontalHeaderLabels(["Timestamp", "Identity", "Image Path"])
        self.log_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.log_table.setEditTriggers(QTableWidget.NoEditTriggers)

        # Layout kiri
        left_layout = QVBoxLayout()
        left_layout.addWidget(self.video_label)
        left_layout.addWidget(self.status_label)
        left_layout.addWidget(self.add_button)
        left_layout.addWidget(self.refresh_button)
        left_layout.addWidget(self.list_button)
        left_layout.addWidget(self.delete_button)
        left_layout.addWidget(self.export_button)

        # Layout kanan
        right_layout = QVBoxLayout()
        right_layout.addWidget(QLabel("Log Deteksi:"))
        right_layout.addWidget(self.log_table)

        # Layout utama
        main_layout = QHBoxLayout()
        main_layout.addLayout(left_layout)
        main_layout.addLayout(right_layout)
        self.setLayout(main_layout)

        # Video capture
        self.cap = cv2.VideoCapture(0)
        if self.cap.isOpened():
            self.status_label.setText("Status: Camera Started")
        else:
            self.status_label.setText("Status: Failed to open camera")

        # Serial ke ESP32
        try:
            self.esp = serial.Serial('COM3', 115200, timeout=1)  # Ganti dengan port ESP32 kamu
            time.sleep(2)  # Tunggu ESP32 siap
            self.status_label.setText("ESP32 Connected")
        except Exception as e:
            self.esp = None
            self.status_label.setText(f"ESP32 Error: {e}")


        self.last_blink_time = time.time()
        self.last_logged_time = 0

        # Logging setup
        self.capture_dir = "capture_logs"
        os.makedirs(self.capture_dir, exist_ok=True)

        self.log_path = "face_recognition_log.xlsx"
        if not os.path.exists(self.log_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["Timestamp", "Identity", "Image Path"])
            wb.save(self.log_path)

        self.load_log_to_table()

        # Timer
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_frame)
        self.timer.start(30)


    def delete_person(self):
        if not os.path.exists(base_dataset_dir):
            self.status_label.setText("Dataset folder tidak ada.")
            return

        people = [d for d in os.listdir(base_dataset_dir) if os.path.isdir(os.path.join(base_dataset_dir, d))]
        if not people:
            self.status_label.setText("Tidak ada data orang untuk dihapus.")
            return

        name, ok = QInputDialog.getItem(self, "Hapus Data", "Pilih orang yang akan dihapus:", people, 0, False)
        if not ok or not name:
            return

        dir_path = os.path.join(base_dataset_dir, name)
        for f in os.listdir(dir_path):
            os.remove(os.path.join(dir_path, f))
        os.rmdir(dir_path)

        self.status_label.setText(f"Data '{name}' dihapus. Klik Refresh Model.")


    def export_log_csv(self):
        if not os.path.exists(self.log_path):
            self.status_label.setText("Log file belum ada.")
            return

        import csv
        from openpyxl import load_workbook

        csv_path = "face_recognition_log.csv"
        wb = load_workbook(self.log_path)
        ws = wb.active

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        self.status_label.setText(f"Log diexport ke {csv_path}")

    def load_log_to_table(self):
        if not os.path.exists(self.log_path):
            self.log_table.setRowCount(0)
            return

        wb = load_workbook(self.log_path)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.log_table.setRowCount(len(rows))

        for i, (timestamp, identity, img_path) in enumerate(rows):
            self.log_table.setItem(i, 0, QTableWidgetItem(str(timestamp)))
            self.log_table.setItem(i, 1, QTableWidgetItem(str(identity)))
            self.log_table.setItem(i, 2, QTableWidgetItem(str(img_path)))

    def list_people(self):
        if not os.path.exists(base_dataset_dir):
            self.status_label.setText("Dataset folder tidak ada.")
            return
        people = [d for d in os.listdir(base_dataset_dir) if os.path.isdir(os.path.join(base_dataset_dir, d))]
        if not people:
            self.status_label.setText("Tidak ada data orang.")
        else:
            self.status_label.setText(f"Data: {', '.join(people)}")

    def update_frame(self):
        ret, frame = self.cap.read()
        if not ret:
            return

        frame = cv2.resize(frame, (640, 480))
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5)

        spoofing = False

        for (x, y, w, h) in faces:
            cv2.rectangle(frame, (x, y), (x+w, y+h), (0,255,0), 2)

            face_roi = frame[y:y+h, x:x+w]
            if face_roi.size == 0:
                continue

            # Recognition
            face_rgb = cv2.cvtColor(face_roi, cv2.COLOR_BGR2RGB)
            face_resized = cv2.resize(face_rgb, (160,160))
            face_tensor = torch.tensor(face_resized).permute(2,0,1).float()/255.0
            face_tensor = (face_tensor - 0.5)/0.5
            face_tensor = face_tensor.unsqueeze(0).to(device)

            with torch.no_grad():
                embedding = resnet(face_tensor).cpu().numpy()[0]
                embedding = embedding / np.linalg.norm(embedding)

            max_similarity = -1
            identity = "Unknown"
            for db_emb, name in zip(db_embeddings, db_names):
                similarity = 1 - cosine(embedding, db_emb)
                if similarity > max_similarity:
                    max_similarity = similarity
                    identity = name

            if max_similarity < 0.7:
                identity = "Unknown"

            # FaceMesh EAR detection
            results = face_mesh.process(face_rgb)
            if results.multi_face_landmarks:
                for face_landmarks in results.multi_face_landmarks:
                    ear_left = get_ear(face_landmarks.landmark, LEFT_EYE, w, h)
                    ear_right = get_ear(face_landmarks.landmark, RIGHT_EYE, w, h)
                    avg_ear = (ear_left + ear_right) / 2.0

                    if avg_ear < EAR_THRESHOLD:
                        self.last_blink_time = time.time()

            if time.time() - self.last_blink_time > 3:
                spoofing = True

            label = f"{identity} ({max_similarity:.2f})"
            if spoofing:
                label += " [FOTO TERDETEKSI]"
                cv2.putText(frame, "FOTO TERDETEKSI", (30,50), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0,0,255), 3)
                self.send_to_esp32(f"FOTO")

            cv2.putText(frame, label, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255,255,0), 2)

            # Log if recognized and not spoofing
            # Inisialisasi dict jika belum ada
            if not hasattr(self, "identity_start_times"):
                self.identity_start_times = {}
                self.identity_last_seen_times = {}
                self.identity_last_logged_times = {}

            now = time.time()

            if identity != "Unknown" and not spoofing:
                # Update waktu terakhir terlihat
                self.identity_last_seen_times[identity] = now

                # Jika identity belum pernah muncul, mulai timer
                if identity not in self.identity_start_times:
                    self.identity_start_times[identity] = now

                # Cek apakah sudah muncul 4 detik
                duration = now - self.identity_start_times[identity]

                # Ambil waktu log terakhir, default 0 jika belum pernah log
                last_log = self.identity_last_logged_times.get(identity, 0)

                if duration >= 4 and (now - last_log >= 10):
                    self.log_detection(identity, frame)
                    self.identity_last_logged_times[identity] = now
                    self.send_to_esp32(f"UNLOCK:{identity}")  # Kirim ke ESP32

            # Reset timer jika wajah hilang lebih dari 1 detik
            to_delete = []
            for ident, last_seen in self.identity_last_seen_times.items():
                if now - last_seen > 1:
                    to_delete.append(ident)

            for ident in to_delete:
                self.identity_start_times.pop(ident, None)
                self.identity_last_seen_times.pop(ident, None)


        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        qt_image = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
        self.video_label.setPixmap(QPixmap.fromImage(qt_image))

    def log_detection(self, identity, frame):
        timestamp = QDateTime.currentDateTime().toString("yyyy-MM-dd_HH-mm-ss")
        filename = f"{identity}_{timestamp}.jpg"
        filepath = os.path.join(self.capture_dir, filename)

        cv2.imwrite(filepath, frame)

        wb = load_workbook(self.log_path)
        ws = wb.active
        ws.append([timestamp.replace("_", " "), identity, filepath])
        wb.save(self.log_path)

        self.status_label.setText(f"Logged: {identity}")
        self.load_log_to_table()

    def add_data(self):
        name, ok = QInputDialog.getText(self, "Add New Face", "Enter Name:")
        if not ok or not name:
            return

        self.status_label.setText("Recording 10 seconds...")
        QApplication.processEvents()

        save_dir = os.path.join(base_dataset_dir, name)
        os.makedirs(save_dir, exist_ok=True)

        existing_files = [f for f in os.listdir(save_dir) if f.lower().endswith(".jpg")]
        indices = []
        for f in existing_files:
            parts = f.split("_")
            if len(parts) == 2 and parts[1].endswith(".jpg"):
                idx_str = parts[1].split(".")[0]
                try:
                    idx = int(idx_str)
                    indices.append(idx)
                except ValueError:
                    continue
        next_idx = max(indices) + 1 if indices else 0

        start_time = QDateTime.currentDateTime()
        end_time = start_time.addSecs(10)

        while QDateTime.currentDateTime() < end_time:
            ret, frame = self.cap.read()
            if not ret:
                continue

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5)

            for (x, y, w, h) in faces:
                face_roi = frame[y:y+h, x:x+w]
                if face_roi.size == 0:
                    continue

                filename = os.path.join(save_dir, f"{name}_{next_idx:03d}.jpg")
                next_idx += 1
                cv2.imwrite(filename, face_roi)

            QApplication.processEvents()

        self.status_label.setText("Recording done. Click 'Refresh Model' to update.")

    def refresh_model(self):
        self.status_label.setText("Refreshing model...")
        QApplication.processEvents()

        all_embeddings = []
        all_names = []

        for person in os.listdir(base_dataset_dir):
            person_dir = os.path.join(base_dataset_dir, person)
            if not os.path.isdir(person_dir):
                continue

            person_embeddings = []

            for file in os.listdir(person_dir):
                if not file.lower().endswith(".jpg"):
                    continue
                img_path = os.path.join(person_dir, file)
                img = cv2.imread(img_path)
                if img is None:
                    continue

                face_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                face_resized = cv2.resize(face_rgb, (160,160))
                face_tensor = torch.tensor(face_resized).permute(2,0,1).float()/255.0
                face_tensor = (face_tensor - 0.5)/0.5
                face_tensor = face_tensor.unsqueeze(0).to(device)

                with torch.no_grad():
                    emb = resnet(face_tensor).cpu().numpy()[0]
                    emb = emb / np.linalg.norm(emb)
                    person_embeddings.append(emb)

            if person_embeddings:
                mean_emb = np.mean(person_embeddings, axis=0)
                all_embeddings.append(mean_emb)
                all_names.append(person)

        if all_embeddings:
            global db_embeddings, db_names
            db_embeddings = np.vstack(all_embeddings)
            db_names = np.array(all_names)

            with open(db_path, "wb") as f:
                pickle.dump({"embeddings": db_embeddings, "names": db_names}, f)

            self.status_label.setText("Model refreshed successfully.")
        else:
            self.status_label.setText("No embeddings found in dataset.")

    def closeEvent(self, event):
        if self.cap:
            self.cap.release()
        event.accept()
    
    def send_to_esp32(self, message):
        if self.esp and self.esp.is_open:
            try:
                self.esp.write((message + "\n").encode())
            except Exception as e:
                self.status_label.setText(f"ESP32 Write Error: {e}")


# ===============================
# Run Application
# ===============================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FaceRecogApp()
    window.show()
    sys.exit(app.exec_())
